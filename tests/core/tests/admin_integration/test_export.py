from datetime import date, datetime
from io import BytesIO
from unittest import mock
from unittest.mock import MagicMock, PropertyMock, patch
from zoneinfo import ZoneInfo

import chardet
import tablib
from core.admin import BookAdmin, BookResource, EBookResource
from core.models import Author, Book, EBook, UUIDCategory
from core.tests.admin_integration.mixins import AdminTestMixin
from core.tests.utils import ignore_utcnow_deprecation_warning
from django import forms
from django.contrib.admin.sites import AdminSite
from django.contrib.admin.views.main import ChangeList
from django.contrib.auth.models import User
from django.core.exceptions import FieldError
from django.http import HttpRequest
from django.test import RequestFactory
from django.test.testcases import TestCase
from django.test.utils import override_settings
from openpyxl.reader.excel import load_workbook
from tablib import Dataset

from import_export import fields, formats, resources, widgets
from import_export.admin import ExportActionMixin, ExportMixin
from import_export.fields import Field
from import_export.formats.base_formats import XLSX
from import_export.resources import ModelResource


class ExportAdminIntegrationTest(AdminTestMixin, TestCase):
    def setUp(self) -> None:
        super().setUp()
        self.bookresource_export_fields_payload = {
            "bookresource_id": True,
            "bookresource_name": True,
            "bookresource_author_email": True,
            "bookresource_categories": True,
        }

    def test_export(self):
        response = self._get_url_response(self.book_export_url)
        self.assertNotIn("Export 0 selected items.", response.content.decode())
        form = response.context["form"]
        self.assertEqual(2, len(form.fields["resource"].choices))

        data = {"format": "0", **self.bookresource_export_fields_payload}
        date_str = datetime.now().strftime("%Y-%m-%d")
        # Should not contain COUNT queries from ModelAdmin.get_results()
        with self.assertNumQueries(5):
            response = self._post_url_response(self.book_export_url, data)
        self.assertTrue(response.has_header("Content-Disposition"))
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertEqual(
            response["Content-Disposition"],
            f'attachment; filename="Book-{date_str}.csv"',
        )
        self.assertEqual(
            b"id,name,author_email,categories\r\n",
            response.content,
        )

    def test_export_with_skip_export_form_from_action(self):
        # setting should have no effect
        with patch(
            "core.admin.BookAdmin.skip_export_form_from_action",
            new_callable=PropertyMock,
            return_value=True,
        ):
            response = self._get_url_response(self.book_export_url)
            target_re = r"This exporter will export the following fields:"
            self.assertRegex(response.content.decode(), target_re)

    @override_settings(IMPORT_EXPORT_SKIP_ADMIN_ACTION_EXPORT_UI=True)
    def test_export_with_skip_export_form_from_action_setting(self):
        # setting should have no effect
        response = self._get_url_response(self.book_export_url)
        target_re = r"This exporter will export the following fields:"
        self.assertRegex(response.content.decode(), target_re)

    @mock.patch("core.admin.BookAdmin.get_export_resource_kwargs")
    def test_export_passes_export_resource_kwargs(
        self, mock_get_export_resource_kwargs
    ):
        # issue 1738
        mock_get_export_resource_kwargs.return_value = {"a": 1}
        self._get_url_response(self.book_export_url)
        self.assertEqual(2, mock_get_export_resource_kwargs.call_count)

    def book_resource_init(self, **kwargs):
        # stub call to the resource constructor
        pass

    @mock.patch.object(BookResource, "__init__", book_resource_init)
    def test_export_passes_no_resource_constructor_params(self):
        # issue 1716
        # assert that the export call with a no-arg constructor
        # does not crash
        self._get_url_response(self.book_export_url)

    def test_get_export_queryset(self):
        model_admin = BookAdmin(Book, AdminSite())

        factory = RequestFactory()
        request = factory.get(self.book_export_url)
        request.user = User.objects.create_user("admin1")

        call_number = 0

        class MyChangeList(ChangeList):
            def get_queryset(self, request):
                nonlocal call_number
                call_number += 1
                return super().get_queryset(request)

        model_admin.get_changelist = lambda request: MyChangeList

        with patch.object(model_admin, "get_paginator") as mock_get_paginator:
            with self.assertNumQueries(4):
                queryset = model_admin.get_export_queryset(request)

            mock_get_paginator.assert_not_called()
            self.assertEqual(call_number, 1)

        self.assertEqual(queryset.count(), Book.objects.count())

    def test_get_export_queryset_no_queryset_init(self):
        """Test if user has own ChangeList which doesn't store queryset during init"""
        model_admin = BookAdmin(Book, AdminSite())

        factory = RequestFactory()
        request = factory.get(self.book_export_url)
        request.user = User.objects.create_user("admin1")

        call_number = 0

        class MyChangeList(ChangeList):
            def __init__(self, *args, **kwargs):
                self.filter_params = {}
                self.model_admin = kwargs.pop("model_admin")
                self.list_filter = kwargs.pop("list_filter")
                self.model = kwargs.pop("model")
                self.date_hierarchy = kwargs.pop("date_hierarchy")
                self.root_queryset = self.model_admin.get_queryset(request)
                self.list_select_related = kwargs.pop("list_select_related")
                self.list_display = kwargs.pop("list_display")
                self.lookup_opts = self.model._meta
                self.params = {}
                self.query = ""

            def get_queryset(self, request):
                nonlocal call_number
                call_number += 1
                return super().get_queryset(request)

        model_admin.get_changelist = lambda request: MyChangeList

        with patch.object(model_admin, "get_paginator") as mock_get_paginator:
            with self.assertNumQueries(4):
                queryset = model_admin.get_export_queryset(request)

            mock_get_paginator.assert_not_called()
            self.assertEqual(call_number, 1)

        self.assertEqual(queryset.count(), Book.objects.count())

    def test_get_export_form_single_resource(self):
        response = self._get_url_response(self.category_export_url)
        content = response.content.decode()
        self.assertNotIn("Export 0 selected items.", content)
        form = response.context["form"]
        self.assertIsInstance(form.fields["resource"].widget, forms.HiddenInput)
        self.assertEqual(form.initial["resource"], "0")

    def test_get_export_FieldError(self):
        # issue 1723
        with mock.patch("import_export.resources.Resource.export") as mock_export:
            mock_export.side_effect = FieldError("some unknown error")
            data = {
                "format": "0",
                "resource": 1,
                "booknameresource_id": True,
                "booknameresource_name": True,
            }
            response = self._post_url_response(self.book_export_url, data)
        target_msg = "Some unknown error"
        self.assertIn(target_msg, response.content.decode())

    def test_export_second_resource(self):
        self._get_url_response(
            self.book_export_url, str_in_response="Export/Import only book names"
        )

        data = {
            "format": "0",
            "resource": 1,
            # Second resource is `BookNameResource`
            "booknameresource_id": True,
            "booknameresource_name": True,
        }
        date_str = datetime.now().strftime("%Y-%m-%d")
        response = self._post_url_response(self.book_export_url, data)
        self.assertTrue(response.has_header("Content-Disposition"))
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertEqual(
            response["Content-Disposition"],
            f'attachment; filename="Book-{date_str}.csv"',
        )
        self.assertEqual(b"id,name\r\n", response.content)

    def test_export_displays_resources_fields(self):
        response = self._get_url_response(self.book_export_url)
        self.assertEqual(
            response.context["fields_list"],
            [
                (
                    "BookResource",
                    [
                        "id",
                        "name",
                        "author",
                        "author_email",
                        "imported",
                        "published",
                        "published_time",
                        "price",
                        "added",
                        "categories",
                    ],
                ),
                ("Export/Import only book names", ["id", "name"]),
            ],
        )

    @override_settings(EXPORT_FORMATS=[XLSX])
    def test_get_export_form_single_format(self):
        response = self._get_url_response(self.category_export_url)
        form = response.context["form"]
        self.assertEqual(1, len(form.fields["format"].choices))
        self.assertTrue(form.fields["format"].widget.attrs["readonly"])
        content = response.content.decode()
        self.assertIn("xlsx", content)
        self.assertNotIn('select name="format"', content)

    @override_settings(EXPORT_FORMATS=[])
    def test_export_empty_export_formats(self):
        with self.assertRaisesRegex(ValueError, "invalid formats list"):
            self._get_url_response(self.category_export_url)

    def test_returns_xlsx_export(self):
        response = self._get_url_response(self.book_export_url)
        self.assertEqual(response.status_code, 200)

        xlsx_index = self._get_input_format_index("xlsx")
        data = {"format": str(xlsx_index), **self.bookresource_export_fields_payload}
        response = self._post_url_response(self.book_export_url, data)
        self.assertTrue(response.has_header("Content-Disposition"))
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @ignore_utcnow_deprecation_warning
    @override_settings(IMPORT_EXPORT_ESCAPE_FORMULAE_ON_EXPORT=True)
    def test_export_escape_formulae(self):
        Book.objects.create(id=1, name="=SUM(1+1)")
        Book.objects.create(id=2, name="<script>alert(1)</script>")
        self._get_url_response(self.book_export_url)

        xlsx_index = self._get_input_format_index("xlsx")
        data = {"format": str(xlsx_index), **self.bookresource_export_fields_payload}
        response = self._post_url_response(self.book_export_url, data)
        content = response.content
        wb = load_workbook(filename=BytesIO(content))
        self.assertEqual("<script>alert(1)</script>", wb.active["B2"].value)
        self.assertEqual("SUM(1+1)", wb.active["B3"].value)

    @override_settings(IMPORT_EXPORT_ESCAPE_FORMULAE_ON_EXPORT=True)
    def test_export_escape_formulae_csv(self):
        b1 = Book.objects.create(id=1, name="=SUM(1+1)")
        self._get_url_response(self.book_export_url)

        index = self._get_input_format_index("csv")
        data = {
            "format": str(index),
            "bookresource_id": True,
            "bookresource_name": True,
        }
        response = self._post_url_response(self.book_export_url, data)
        self.assertIn(
            f"{b1.id},SUM(1+1)\r\n".encode(),
            response.content,
        )

    @override_settings(IMPORT_EXPORT_ESCAPE_FORMULAE_ON_EXPORT=False)
    def test_export_escape_formulae_csv_false(self):
        b1 = Book.objects.create(id=1, name="=SUM(1+1)")
        self._get_url_response(self.book_export_url)

        index = self._get_input_format_index("csv")
        data = {
            "format": str(index),
            "bookresource_id": True,
            "bookresource_name": True,
        }
        response = self._post_url_response(self.book_export_url, data)
        self.assertIn(
            f"{b1.id},=SUM(1+1)\r\n".encode(),
            response.content,
        )

    def test_export_model_with_custom_PK(self):
        # issue 1800
        UUIDCategory.objects.create(name="UUIDCategory")
        response = self._get_url_response(self.uuid_category_export_url)
        form = response.context["form"]
        self.assertEqual(
            form.fields["resource"].choices,
            [(0, "UUIDCategoryResource")],
        )

    def test_export_get(self):
        """
        Test export view get method.
        Test that field checkboxes are displayed with names as discussed under #1846
        """
        response = self._get_url_response(self.ebook_export_url)
        self.assertContains(
            response,
            '<label for="id_ebookresource_published">'
            "Published (published_date)</label>",
            html=True,
        )
        self.assertContains(
            response,
            '<input type="checkbox" name="ebookresource_published" resource-id="0" '
            'id="id_ebookresource_published" checked="">',
            html=True,
        )

    def test_export_with_custom_field(self):
        # issue 1808
        a = Author.objects.create(id=11, name="Ian Fleming")
        data = {
            "format": "0",
            "author": a.id,
            "resource": "",
            "ebookresource_id": True,
            "ebookresource_author_email": True,
            "ebookresource_name": True,
            "ebookresource_published": True,
        }
        date_str = datetime.now().strftime("%Y-%m-%d")
        response = self._post_url_response(self.ebook_export_url, data)
        self.assertTrue(response.has_header("Content-Disposition"))
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertEqual(
            response["Content-Disposition"],
            f'attachment; filename="EBook-{date_str}.csv"',
        )
        self.assertEqual(
            b"id,Email of the author,name,published_date\r\n", response.content
        )


class FilteredExportAdminIntegrationTest(AdminTestMixin, TestCase):
    fixtures = ["category", "book", "author"]

    def test_export_filters_by_form_param(self):
        # issue 1578
        author = Author.objects.get(name="Ian Fleming")

        data = {
            "format": "0",
            "author": str(author.id),
            "ebookresource_id": True,
            "ebookresource_author_email": True,
            "ebookresource_name": True,
            "ebookresource_published": True,
        }
        date_str = datetime.now().strftime("%Y-%m-%d")
        response = self._post_url_response(self.ebook_export_url, data)
        self.assertTrue(response.has_header("Content-Disposition"))
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertEqual(
            response["Content-Disposition"],
            f'attachment; filename="EBook-{date_str}.csv"',
        )
        self.assertEqual(
            b"id,Email of the author,name,published_date\r\n"
            b"5,ian@example.com,The Man with the Golden Gun,1965-04-01\r\n",
            response.content,
        )


class TestExportEncoding(TestCase):
    mock_request = MagicMock(spec=HttpRequest)
    mock_request.POST = {"format": 0, "bookresource_id": True}

    class TestMixin(ExportMixin):
        model = Book

        def __init__(self, test_str=None):
            self.test_str = test_str

        def get_data_for_export(self, request, queryset, **kwargs):
            dataset = Dataset(headers=["id", "name"])
            dataset.append([1, self.test_str])
            return dataset

        def get_export_queryset(self, request):
            return []

        def get_export_filename(self, request, queryset, file_format):
            return "f"

    def setUp(self):
        self.file_format = formats.base_formats.CSV()
        self.export_mixin = self.TestMixin(test_str="teststr")

    def test_to_encoding_not_set_default_encoding_is_utf8(self):
        self.export_mixin = self.TestMixin(test_str="teststr")
        data = self.export_mixin.get_export_data(
            self.file_format, self.mock_request, []
        )
        csv_dataset = tablib.import_set(data)
        self.assertEqual("teststr", csv_dataset.dict[0]["name"])

    def test_to_encoding_set(self):
        self.export_mixin = self.TestMixin(test_str="ハローワールド")
        data = self.export_mixin.get_export_data(
            self.file_format, self.mock_request, [], encoding="shift-jis"
        )
        encoding = chardet.detect(bytes(data))["encoding"]
        self.assertEqual("SHIFT_JIS", encoding)

    def test_to_encoding_set_incorrect(self):
        self.export_mixin = self.TestMixin()
        with self.assertRaises(LookupError):
            self.export_mixin.get_export_data(
                self.file_format,
                self.mock_request,
                [],
                encoding="bad-encoding",
            )

    @ignore_utcnow_deprecation_warning
    def test_to_encoding_not_set_for_binary_file(self):
        self.export_mixin = self.TestMixin(test_str="teststr")
        self.file_format = formats.base_formats.XLSX()
        data = self.export_mixin.get_export_data(
            self.file_format,
            self.mock_request,
            [],
        )
        binary_dataset = tablib.import_set(data)
        self.assertEqual("teststr", binary_dataset.dict[0]["name"])

    def test_export_action_to_encoding(self):
        self.export_mixin.to_encoding = "utf-8"
        with mock.patch(
            "import_export.admin.ExportMixin.get_export_data"
        ) as mock_get_export_data:
            self.export_mixin.export_action(self.mock_request)
            encoding_kwarg = mock_get_export_data.call_args_list[0][1]["encoding"]
            self.assertEqual("utf-8", encoding_kwarg)

    @override_settings(IMPORT_EXPORT_SKIP_ADMIN_ACTION_EXPORT_UI=True)
    def test_export_admin_action_to_encoding(self):
        class TestExportActionMixin(ExportActionMixin):
            def get_export_filename(self, request, queryset, file_format):
                return "f"

        self.export_mixin = TestExportActionMixin()
        self.export_mixin.to_encoding = "utf-8"
        with mock.patch(
            "import_export.admin.ExportMixin.get_export_data"
        ) as mock_get_export_data:
            self.export_mixin.export_admin_action(self.mock_request, [])
            encoding_kwarg = mock_get_export_data.call_args_list[0][1]["encoding"]
            self.assertEqual("utf-8", encoding_kwarg)


class TestSelectableFieldsExportPage(AdminTestMixin, TestCase):
    def test_selectable_fields_rendered_with_resource_index_attribute(self) -> None:
        response = self._get_url_response(self.book_export_url)
        form_resources = response.context["form"].resources
        content = response.content.decode()
        for index, resource in enumerate(form_resources):
            resource_fields = resource().get_export_order()
            self.assertEqual(
                content.count(f'resource-index="{index}"'),
                len(resource_fields),
            )


class CustomColumnNameExportTest(AdminTestMixin, TestCase):
    """Test export ok when column name is defined in fields list (issue 1828)."""

    def setUp(self):
        super().setUp()
        self.author = Author.objects.create(id=11, name="Ian Fleming")
        self.book = Book.objects.create(
            name="Moonraker", author=self.author, published=date(1955, 4, 5)
        )
        self.orig_fields = EBookResource._meta.fields
        EBookResource._meta.fields = (
            "id",
            "author_email",
            "name",
            "published_date",
            "auteur_name",
        )

    def tearDown(self):
        super().tearDown()
        EBookResource._meta.fields = self.orig_fields

    def test_export_with_custom_field(self):
        data = {
            "format": "0",
            "author": self.author.id,
            "resource": "",
            "ebookresource_id": True,
            "ebookresource_author_email": True,
            "ebookresource_name": True,
            "ebookresource_published_date": True,
        }
        date_str = datetime.now().strftime("%Y-%m-%d")
        response = self._post_url_response(self.ebook_export_url, data)
        self.assertTrue(response.has_header("Content-Disposition"))
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertEqual(
            response["Content-Disposition"],
            f'attachment; filename="EBook-{date_str}.csv"',
        )
        s = (
            "id,Email of the author,name,published_date\r\n"
            f"{self.book.id},,Moonraker,1955-04-05\r\n"
        )
        self.assertEqual(s, response.content.decode())

    def test_export_with_custom_name(self):
        # issue 1893
        data = {
            "format": "0",
            "author": self.author.id,
            "resource": "",
            "ebookresource_id": True,
            "ebookresource_author_email": True,
            "ebookresource_name": True,
            "ebookresource_published_date": True,
            "ebookresource_auteur_name": True,
        }
        response = self._post_url_response(self.ebook_export_url, data)
        s = (
            "id,Email of the author,name,published_date,Author Name\r\n"
            f"{self.book.id},,Moonraker,1955-04-05,Ian Fleming\r\n"
        )
        self.assertEqual(s, response.content.decode())


class DeclaredFieldWithAttributeExportTest(AdminTestMixin, TestCase):
    """
    If a custom field is declared, export should work
    even if no `fields` declaration is present.
    (issue 1953)
    """

    class _BookResource(ModelResource):
        name = Field(attribute="author__name", column_name="Author Name")

        class Meta:
            model = Book

    def setUp(self):
        super().setUp()
        self.author = Author.objects.create(id=11, name="Ian Fleming")
        self.book = Book.objects.create(
            name="Moonraker", author=self.author, published=date(1955, 4, 5)
        )

    @patch("import_export.mixins.BaseExportMixin.choose_export_resource_class")
    def test_export_with_declared_author_name_field(
        self, mock_choose_export_resource_class
    ):
        mock_choose_export_resource_class.return_value = self._BookResource
        data = {
            "format": "0",
            "resource": "0",
            "bookresource_name": True,
        }
        response = self._post_url_response(self.book_export_url, data)
        s = "Author Name\r\nIan Fleming\r\n"
        self.assertEqual(s, response.content.decode())


class DeclaredFieldWithAttributeAndFieldsExportTest(AdminTestMixin, TestCase):
    """
    If a custom field is declared, export should work
    when `fields` declaration is present.
    (issue 1953)
    """

    class _BookResource(ModelResource):
        name = Field(attribute="author__name", column_name="Author Name")

        class Meta:
            fields = ("name",)
            model = Book

    def setUp(self):
        super().setUp()
        self.author = Author.objects.create(id=11, name="Ian Fleming")
        self.book = Book.objects.create(
            name="Moonraker", author=self.author, published=date(1955, 4, 5)
        )

    @patch("import_export.mixins.BaseExportMixin.choose_export_resource_class")
    def test_export_with_declared_author_name_field(
        self, mock_choose_export_resource_class
    ):
        mock_choose_export_resource_class.return_value = self._BookResource
        data = {
            "format": "0",
            "resource": "0",
            "bookresource_name": True,
        }
        response = self._post_url_response(self.book_export_url, data)
        s = "Author Name\r\nIan Fleming\r\n"
        self.assertEqual(s, response.content.decode())


class DeclaredFieldWithNoAttributeExportTest(AdminTestMixin, TestCase):
    """
    If a custom field is declared with no attribute the field will be present
    but with an empty string.
    """

    class _BookResource(ModelResource):
        author_email = Field(column_name="Author Email")

        class Meta:
            model = Book

    def setUp(self):
        super().setUp()
        self.book = Book.objects.create(
            name="Moonraker", author_email="ian@fleming.com"
        )

    @patch("import_export.mixins.BaseExportMixin.choose_export_resource_class")
    def test_export_with_declared_author_email_field(
        self, mock_choose_export_resource_class
    ):
        mock_choose_export_resource_class.return_value = self._BookResource
        data = {"format": "0", "resource": "0", "bookresource_author_email": True}
        response = self._post_url_response(self.book_export_url, data)
        s = 'Author Email\r\n""\r\n'
        self.assertEqual(s, response.content.decode())


class DeclaredFieldWithIncorrectNameInFieldsExportTest(AdminTestMixin, TestCase):
    """
    If a custom field is declared with no attribute the process should not crash
    if that field is not in `fields`.
    issue #1959
    """

    def setUp(self):
        super().setUp()
        self.author = Author.objects.create(id=11, name="Ian Fleming")
        self.book = Book.objects.create(
            name="Moonraker", author_email="ian@fleming.com", author=self.author
        )
        self.orig_fields = EBookResource._meta.fields
        EBookResource._meta.fields = ("a",)

    def tearDown(self):
        super().tearDown()
        EBookResource._meta.fields = self.orig_fields

    def test_export_with_declared_author_email_field(self):
        data = {
            "format": "0",
            "resource": "0",
            "ebookresource_id": True,
            "ebookresource_a": True,
            "author": self.author.id,
        }
        with self.assertWarns(UserWarning) as w:
            response = self._post_url_response(self.ebook_export_url, data)
            self.assertEqual(
                "cannot identify field for export with name 'a'",
                str(w.warnings[-1].message),
            )
        s = f"id\r\n{self.book.id}\r\n"
        self.assertEqual(s, response.content.decode())


class FilteredExportTest(AdminTestMixin, TestCase):
    """
    Tests that exports can be filtered by a custom form field.
    This process is demonstrated in the documentation.
    """

    def test_filtered_export(self):
        a1 = Author.objects.create(id=11, name="Ian Fleming")
        a2 = Author.objects.create(id=12, name="James Joyce")
        b1 = Book.objects.create(name="Moonraker", author=a1)
        b2 = Book.objects.create(name="Ulysses", author=a2)
        self._get_url_response(self.ebook_export_url)
        data = {
            "format": "0",
            "author": a1.id,
            "resource": "",
            "ebookresource_id": True,
            "ebookresource_name": True,
        }
        response = self._post_url_response(self.ebook_export_url, data)
        s = "id,name\r\n" f"{b1.id},Moonraker\r\n"
        self.assertEqual(s.encode(), response.content)

        data["author"] = a2.id
        response = self._post_url_response(self.ebook_export_url, data)
        s = "id,name\r\n" f"{b2.id},Ulysses\r\n"
        self.assertEqual(s.encode(), response.content)


class SkipExportFormResourceConfigTest(AdminTestMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.model_admin = BookAdmin(EBook, AdminSite())

        book = Book.objects.create(name="Moonraker", published=date(1955, 4, 5))
        self.target_file_contents = (
            "id,name,author,author_email,imported,published,"
            "published_time,price,added,categories\r\n"
            f"{book.id},Moonraker,,,0,1955-04-05,,,,\r\n"
        )

        factory = RequestFactory()
        self.request = factory.get(self.book_export_url, follow=True)
        self.request.user = User.objects.create_user("admin1")

    def test_export_skips_export_form(self):
        self.model_admin.skip_export_form = True
        response = self.model_admin.export_action(self.request)
        self._check_export_file_response(
            response, self.target_file_contents, file_prefix="EBook"
        )

    @override_settings(IMPORT_EXPORT_SKIP_ADMIN_EXPORT_UI=True)
    def test_export_skips_export_form_setting_enabled(self):
        response = self.model_admin.export_action(self.request)
        self._check_export_file_response(
            response, self.target_file_contents, file_prefix="EBook"
        )


class ExportBinaryFieldsTest(AdminTestMixin, TestCase):
    # Test that Dates, Booleans, numbers etc are retained as native types
    # when exporting to XLSX, XLS, ODS (see #1939)

    class DeclaredModelFieldBookResource(resources.ModelResource):
        # declare a field and enforce export output as str (coerce_to_string)
        id = fields.Field(
            attribute="id",
            widget=widgets.NumberWidget(coerce_to_string=True),
        )
        imported = fields.Field(
            attribute="imported",
            widget=widgets.BooleanWidget(coerce_to_string=True),
        )
        published = fields.Field(
            attribute="published",
            widget=widgets.DateWidget("%d.%m.%Y", coerce_to_string=True),
        )

        class Meta:
            model = Book
            export_order = ("id", "imported", "published")

    def test_dynamic_type_export(self):
        Book.objects.create(id=101, published=datetime(2010, 8, 2), imported=True)
        data = {
            "format": "2",
            "bookresource_id": True,
            "bookresource_imported": True,
            "bookresource_published": True,
        }
        response = self.client.post(self.book_export_url, data)
        self.assertEqual(response.status_code, 200)
        content = response.content
        wb = load_workbook(filename=BytesIO(content))
        self.assertEqual(101, wb.active["A2"].value)
        self.assertEqual(True, wb.active["B2"].value)
        self.assertEqual(datetime(2010, 8, 2), wb.active["C2"].value)

    @patch("import_export.mixins.BaseExportMixin.choose_export_resource_class")
    def test_dynamic_export_with_custom_resource(
        self, mock_choose_export_resource_class
    ):
        # Test that `coerce_to_string` is ignored
        mock_choose_export_resource_class.return_value = (
            self.DeclaredModelFieldBookResource
        )
        Book.objects.create(id=101, published=date(2000, 8, 2), imported=True)
        data = {
            "format": "2",
            "bookresource_id": True,
            "bookresource_imported": True,
            "bookresource_published": True,
        }
        response = self.client.post(self.book_export_url, data)
        self.assertEqual(response.status_code, 200)
        content = response.content
        wb = load_workbook(filename=BytesIO(content))
        self.assertEqual(101, wb.active["A2"].value)
        self.assertEqual(1, wb.active["B2"].value)
        self.assertEqual(datetime(2000, 8, 2), wb.active["C2"].value)


@override_settings(USE_TZ=True, TIME_ZONE="UTC")
class ExportTzAwareDateTest(AdminTestMixin, TestCase):
    # issue 1995
    # test that tz aware dates do not crash on export
    class BookResource_(resources.ModelResource):

        class Meta:
            model = Book
            fields = ("id", "name", "added")

    @patch("import_export.mixins.BaseExportMixin.choose_export_resource_class")
    def test_datetime_export_xlsx(self, mock_choose_export_resource_class):
        mock_choose_export_resource_class.return_value = self.BookResource_
        date_added = datetime(2024, 11, 8, 14, 40, tzinfo=ZoneInfo("UTC"))
        Book.objects.create(id=101, name="Moonraker", added=date_added)

        data = {
            "format": "2",
            "bookresource_id": True,
            "bookresource_added": True,
        }
        response = self.client.post(self.book_export_url, data)
        self.assertEqual(response.status_code, 200)
        content = response.content
        wb = load_workbook(filename=BytesIO(content))
        self.assertEqual(date_added.replace(tzinfo=None), wb.active["B2"].value)

    @override_settings(TIME_ZONE="Asia/Hong_Kong")
    @patch("import_export.mixins.BaseExportMixin.choose_export_resource_class")
    def test_datetime_export_xlsx_with_timezone(
        self, mock_choose_export_resource_class
    ):
        mock_choose_export_resource_class.return_value = self.BookResource_
        date_added = datetime(2024, 11, 8, 14, 40, tzinfo=ZoneInfo("Asia/Hong_Kong"))
        Book.objects.create(id=101, name="Moonraker", added=date_added)

        data = {
            "format": "2",
            "bookresource_id": True,
            "bookresource_added": True,
        }
        response = self.client.post(self.book_export_url, data)
        self.assertEqual(response.status_code, 200)
        content = response.content
        wb = load_workbook(filename=BytesIO(content))
        self.assertEqual(date_added.replace(tzinfo=None), wb.active["B2"].value)

    @patch("import_export.mixins.BaseExportMixin.choose_export_resource_class")
    def test_datetime_export_xls(self, mock_choose_export_resource_class):
        mock_choose_export_resource_class.return_value = self.BookResource_
        date_added = datetime(2024, 11, 8, 14, 40, tzinfo=ZoneInfo("UTC"))
        Book.objects.create(id=101, name="Moonraker", added=date_added)

        data = {
            "format": "1",
            "bookresource_id": True,
            "bookresource_added": True,
        }
        response = self.client.post(self.book_export_url, data)
        self.assertEqual(response.status_code, 200)

    @patch("import_export.mixins.BaseExportMixin.choose_export_resource_class")
    def test_datetime_export_ods(self, mock_choose_export_resource_class):
        mock_choose_export_resource_class.return_value = self.BookResource_
        date_added = datetime(2024, 11, 8, 14, 40, tzinfo=ZoneInfo("UTC"))
        Book.objects.create(id=101, name="Moonraker", added=date_added)

        data = {
            "format": "4",
            "bookresource_id": True,
            "bookresource_added": True,
        }
        response = self.client.post(self.book_export_url, data)
        self.assertEqual(response.status_code, 200)

    @patch("import_export.mixins.BaseExportMixin.choose_export_resource_class")
    def test_datetime_export_empty_field(self, mock_choose_export_resource_class):
        mock_choose_export_resource_class.return_value = self.BookResource_
        date_added = None
        Book.objects.create(id=101, name="Moonraker", added=date_added)

        data = {
            "format": "2",
            "bookresource_id": True,
            "bookresource_added": True,
        }
        response = self.client.post(self.book_export_url, data)
        self.assertEqual(response.status_code, 200)
        content = response.content
        wb = load_workbook(filename=BytesIO(content))
        self.assertIsNone(wb.active["B2"].value)


class ExportInvalidCharTest(AdminTestMixin, TestCase):
    # issue 2000

    def test_export_xlsx(self):
        Book.objects.create(id=101, name="invalid" + chr(11))

        data = {
            "format": "2",
            "bookresource_id": True,
            "bookresource_name": True,
        }
        response = self.client.post(self.book_export_url, data)
        self.assertIn(
            "Export failed due to IllegalCharacterError", response.content.decode()
        )

    @override_settings(IMPORT_EXPORT_ESCAPE_ILLEGAL_CHARS_ON_EXPORT=True)
    def test_export_xlsx_with_escape(self):
        Book.objects.create(id=101, name="invalid" + chr(11))

        data = {
            "format": "2",
            "bookresource_id": True,
            "bookresource_name": True,
        }
        response = self.client.post(self.book_export_url, data)
        self.assertEqual(response.status_code, 200)
        content = response.content
        wb = load_workbook(filename=BytesIO(content))
        self.assertEqual("invalid�", wb.active["B2"].value)
