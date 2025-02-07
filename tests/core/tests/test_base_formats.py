import os
import unittest
from io import BytesIO
from unittest import mock

import openpyxl
import tablib
from core.tests.utils import ignore_utcnow_deprecation_warning
from django.test import TestCase, override_settings
from django.utils.encoding import force_str
from tablib.core import UnsupportedFormat

from import_export.formats import base_formats
from import_export.widgets import NumberWidget


class FormatTest(TestCase):
    def setUp(self):
        self.format = base_formats.Format()

    @mock.patch(
        "import_export.formats.base_formats.HTML.get_format", side_effect=ImportError
    )
    def test_format_non_available1(self, mocked):
        self.assertFalse(base_formats.HTML.is_available())

    @mock.patch(
        "import_export.formats.base_formats.HTML.get_format",
        side_effect=UnsupportedFormat,
    )
    def test_format_non_available2(self, mocked):
        self.assertFalse(base_formats.HTML.is_available())

    def test_format_available(self):
        self.assertTrue(base_formats.CSV.is_available())

    def test_get_title(self):
        self.assertEqual(
            "<class 'import_export.formats.base_formats.Format'>",
            str(self.format.get_title()),
        )

    def test_create_dataset_NotImplementedError(self):
        with self.assertRaises(NotImplementedError):
            self.format.create_dataset(None)

    def test_export_data_NotImplementedError(self):
        with self.assertRaises(NotImplementedError):
            self.format.export_data(None)

    def test_get_extension(self):
        self.assertEqual("", self.format.get_extension())

    def test_get_content_type(self):
        self.assertEqual("application/octet-stream", self.format.get_content_type())

    def test_is_available_default(self):
        self.assertTrue(self.format.is_available())

    def test_can_import_default(self):
        self.assertFalse(self.format.can_import())

    def test_can_export_default(self):
        self.assertFalse(self.format.can_export())


class TablibFormatTest(TestCase):
    def setUp(self):
        self.format = base_formats.TablibFormat()

    def test_get_format_for_undefined_TABLIB_MODULE_raises_AttributeError(self):
        with self.assertRaises(AttributeError):
            self.format.get_format()


class XLSTest(TestCase):
    def setUp(self):
        self.format = base_formats.XLS()

    def test_binary_format(self):
        self.assertTrue(self.format.is_binary())

    def test_import(self):
        filename = os.path.join(
            os.path.dirname(__file__), os.path.pardir, "exports", "books.xls"
        )
        with open(filename, self.format.get_read_mode()) as in_stream:
            self.format.create_dataset(in_stream.read())


class XLSXTest(TestCase):
    def setUp(self):
        self.format = base_formats.XLSX()
        self.filename = os.path.join(
            os.path.dirname(__file__), os.path.pardir, "exports", "books.xlsx"
        )

    def test_binary_format(self):
        self.assertTrue(self.format.is_binary())

    @ignore_utcnow_deprecation_warning
    def test_import(self):
        with open(self.filename, self.format.get_read_mode()) as in_stream:
            dataset = self.format.create_dataset(in_stream.read())
        result = dataset.dict
        self.assertEqual(1, len(result))
        row = result.pop()
        self.assertEqual(1, row["id"])
        self.assertEqual("Some book", row["name"])
        self.assertEqual("test@example.com", row["author_email"])
        self.assertEqual(4, row["price"])

    @mock.patch("openpyxl.load_workbook")
    def test_that_load_workbook_called_with_required_args(self, mock_load_workbook):
        self.format.create_dataset(b"abc")
        mock_load_workbook.assert_called_with(
            unittest.mock.ANY, read_only=True, data_only=True
        )

    @override_settings(IMPORT_EXPORT_IMPORT_IGNORE_BLANK_LINES=False)
    def test_xlsx_create_dataset__empty_rows(self):
        """Default situation without the flag: do not ignore the empty rows for
        backwards compatibility.
        """
        rows_before = 3
        empty_rows = 5
        rows_after = 2

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2", "Header3"])

        for _ in range(rows_before):
            ws.append(["Data1", "Data2", "Data3"])

        for _ in range(empty_rows):
            ws.append([None, None, None])

        for _ in range(rows_after):
            ws.append(["Data1", "Data2", "Data3"])

        xlsx_data = BytesIO()
        wb.save(xlsx_data)
        xlsx_data.seek(0)

        dataset = self.format.create_dataset(xlsx_data.getvalue())
        assert len(dataset) == rows_before + empty_rows + rows_after  # With empty rows

    @override_settings(IMPORT_EXPORT_IMPORT_IGNORE_BLANK_LINES=True)
    def test_xlsx_create_dataset__ignore_empty_rows(self):
        """Ensure that empty rows are not added to the dataset."""
        rows_before = 3
        empty_rows = 5
        rows_after = 2

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2", "Header3"])

        for _ in range(rows_before):
            ws.append(["Data1", "Data2", "Data3"])

        for _ in range(empty_rows):
            ws.append([None, None, None])

        for _ in range(rows_after):
            ws.append(["Data1", "Data2", "Data3"])

        xlsx_data = BytesIO()
        wb.save(xlsx_data)
        xlsx_data.seek(0)

        dataset = self.format.create_dataset(xlsx_data.getvalue())
        assert len(dataset) == rows_before + rows_after  # Without empty rows


class CSVTest(TestCase):
    def setUp(self):
        self.format = base_formats.CSV()
        self.dataset = tablib.Dataset(headers=["id", "username"])
        self.dataset.append(("1", "x"))

    def test_import_dos(self):
        filename = os.path.join(
            os.path.dirname(__file__), os.path.pardir, "exports", "books-dos.csv"
        )
        with open(filename, self.format.get_read_mode()) as in_stream:
            actual = in_stream.read()
        expected = "id,name,author_email\n1,Some book,test@example.com\n"
        self.assertEqual(actual, expected)

    def test_import_mac(self):
        filename = os.path.join(
            os.path.dirname(__file__), os.path.pardir, "exports", "books-mac.csv"
        )
        with open(filename, self.format.get_read_mode()) as in_stream:
            actual = in_stream.read()
        expected = "id,name,author_email\n1,Some book,test@example.com\n"
        self.assertEqual(actual, expected)

    def test_import_unicode(self):
        # importing csv UnicodeEncodeError 347
        filename = os.path.join(
            os.path.dirname(__file__), os.path.pardir, "exports", "books-unicode.csv"
        )
        with open(filename, self.format.get_read_mode()) as in_stream:
            data = force_str(in_stream.read())
        base_formats.CSV().create_dataset(data)

    def test_export_data(self):
        res = self.format.export_data(self.dataset)
        self.assertEqual("id,username\r\n1,x\r\n", res)

    def test_get_extension(self):
        self.assertEqual("csv", self.format.get_extension())

    def test_content_type(self):
        self.assertEqual("text/csv", self.format.get_content_type())

    def test_can_import(self):
        self.assertTrue(self.format.can_import())

    def test_can_export(self):
        self.assertTrue(self.format.can_export())


class TSVTest(TestCase):
    def setUp(self):
        self.format = base_formats.TSV()

    def test_import_mac(self):
        filename = os.path.join(
            os.path.dirname(__file__), os.path.pardir, "exports", "books-mac.tsv"
        )
        with open(filename, self.format.get_read_mode()) as in_stream:
            actual = in_stream.read()
        expected = "id\tname\tauthor_email\n1\tSome book\ttest@example.com\n"
        self.assertEqual(actual, expected)

    def test_import_unicode(self):
        # importing tsv UnicodeEncodeError
        filename = os.path.join(
            os.path.dirname(__file__), os.path.pardir, "exports", "books-unicode.tsv"
        )
        with open(filename, self.format.get_read_mode()) as in_stream:
            data = force_str(in_stream.read())
        base_formats.TSV().create_dataset(data)


class TextFormatTest(TestCase):
    def setUp(self):
        self.format = base_formats.TextFormat()

    def test_get_read_mode(self):
        self.assertEqual("r", self.format.get_read_mode())

    def test_is_binary(self):
        self.assertFalse(self.format.is_binary())


class HTMLFormatTest(TestCase):
    def setUp(self):
        self.format = base_formats.HTML()
        self.dataset = tablib.Dataset(headers=["id", "username", "name"])
        self.dataset.append((1, "good_user", "John Doe"))
        self.dataset.append(
            (
                "2",
                "evil_user",
                '<script>alert("I want to steal your credit card data")</script>',
            )
        )

    def test_export_html_escape(self):
        res = self.format.export_data(self.dataset)
        self.assertIn(
            (
                "<tr><td>1</td>"
                "<td>good_user</td>"
                "<td>John Doe</td></tr>"
                "<tr><td>2</td>"
                "<td>evil_user</td>"
                '<td>&lt;script&gt;alert("I want to steal your credit card data")'
                "&lt;/script&gt;</td></tr>"
            ),
            res,
        )


class YAMLFormatTest(TestCase):
    def test_numeric_widget_export(self):
        dataset = tablib.Dataset(headers=["id", "username"])
        dataset.append((NumberWidget().render(1), "x"))
        res = base_formats.YAML().export_data(dataset)
        self.assertEqual("- {id: '1', username: x}\n", res)
